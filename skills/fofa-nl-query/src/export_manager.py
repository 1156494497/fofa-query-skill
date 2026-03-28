"""
导出管理器
处理查询结果的导出，支持 Excel 和 CSV 格式
"""

import os
import csv
from typing import Dict, List, Any, Optional
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .fofa_client import FofaResult


class ExportManager:
    """导出管理器"""
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        初始化导出管理器
        
        Args:
            output_dir: 输出目录，默认为当前目录
        """
        self.output_dir = output_dir or os.getcwd()
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)
    
    def export_excel(self, result: FofaResult, filename: Optional[str] = None,
                     include_summary: bool = True) -> str:
        """
        导出为 Excel 格式
        
        Args:
            result: Fofa 查询结果
            filename: 输出文件名，默认为自动生成
            include_summary: 是否包含摘要工作表
            
        Returns:
            str: 输出文件路径
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"fofa_query_results_{timestamp}.xlsx"
        
        filepath = os.path.join(self.output_dir, filename)
        
        # 创建工作簿
        wb = Workbook()
        
        # 第一个工作表：查询结果
        ws_results = wb.active
        ws_results.title = "查询结果"
        
        if result.results:
            # 将结果转换为 DataFrame
            df = pd.DataFrame(result.results, columns=result.fields)
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws_results.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 设置表头样式
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="center")
            
            # 调整列宽
            for column in ws_results.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_results.column_dimensions[column_letter].width = adjusted_width
        
        # 第二个工作表：统计摘要
        if include_summary:
            ws_summary = wb.create_sheet("统计摘要")
            self._create_summary_sheet(ws_summary, result)
        
        # 第三个工作表：查询信息
        ws_info = wb.create_sheet("查询信息")
        self._create_info_sheet(ws_info, result)
        
        # 保存文件
        wb.save(filepath)
        
        return filepath
    
    def export_csv(self, result: FofaResult, filename: Optional[str] = None) -> str:
        """
        导出为 CSV 格式
        
        Args:
            result: Fofa 查询结果
            filename: 输出文件名，默认为自动生成
            
        Returns:
            str: 输出文件路径
        """
        if filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"fofa_query_results_{timestamp}.csv"
        
        filepath = os.path.join(self.output_dir, filename)
        
        if result.results:
            df = pd.DataFrame(result.results, columns=result.fields)
            df.to_csv(filepath, index=False, encoding='utf-8-sig')
        else:
            # 创建空文件
            with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow(result.fields if result.fields else [])
        
        return filepath
    
    def export_both(self, result: FofaResult, base_filename: Optional[str] = None) -> Dict[str, str]:
        """
        同时导出 Excel 和 CSV 格式
        
        Args:
            result: Fofa 查询结果
            base_filename: 基础文件名（不含扩展名）
            
        Returns:
            Dict[str, str]: 包含两种格式文件路径的字典
        """
        if base_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_filename = f"fofa_query_results_{timestamp}"
        
        excel_path = self.export_excel(result, f"{base_filename}.xlsx")
        csv_path = self.export_csv(result, f"{base_filename}.csv")
        
        return {
            'excel': excel_path,
            'csv': csv_path
        }
    
    def _create_summary_sheet(self, ws, result: FofaResult):
        """创建统计摘要工作表"""
        row = 1
        
        # 标题
        ws.cell(row=row, column=1, value="查询统计摘要")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        # 基本信息
        ws.cell(row=row, column=1, value="查询语句:")
        ws.cell(row=row, column=2, value=result.query)
        row += 1
        
        ws.cell(row=row, column=1, value="总结果数:")
        ws.cell(row=row, column=2, value=result.total)
        row += 1
        
        ws.cell(row=row, column=1, value="返回结果数:")
        ws.cell(row=row, column=2, value=len(result.results))
        row += 2
        
        if not result.results:
            ws.cell(row=row, column=1, value="无数据")
            return
        
        # 按国家分布
        row = self._add_distribution_stats(ws, result, 'country', '国家分布', row)
        
        # 按省份分布
        row = self._add_distribution_stats(ws, result, 'region', '省份/地区分布', row)
        
        # 按端口分布
        row = self._add_distribution_stats(ws, result, 'port', '端口分布', row)
        
        # 按协议分布
        row = self._add_distribution_stats(ws, result, 'protocol', '协议分布', row)
        
        # 按服务器类型分布
        row = self._add_distribution_stats(ws, result, 'server', '服务器类型分布', row)
    
    def _add_distribution_stats(self, ws, result: FofaResult, field: str, 
                                title: str, start_row: int) -> int:
        """添加分布统计"""
        row = start_row
        
        # 检查字段是否存在
        if field not in result.fields:
            return row
        
        # 统计分布
        distribution = {}
        for item in result.results:
            value = item.get(field, '未知')
            if value:
                distribution[value] = distribution.get(value, 0) + 1
        
        if not distribution:
            return row
        
        # 标题
        ws.cell(row=row, column=1, value=title)
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        # 表头
        ws.cell(row=row, column=1, value=field.capitalize())
        ws.cell(row=row, column=2, value="数量")
        ws.cell(row=row, column=3, value="占比")
        for col in range(1, 4):
            ws.cell(row=row, column=col).font = Font(bold=True)
            ws.cell(row=row, column=col).fill = PatternFill(
                start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
            )
        row += 1
        
        # 排序并显示前 10
        sorted_items = sorted(distribution.items(), key=lambda x: x[1], reverse=True)[:10]
        total = len(result.results)
        
        for value, count in sorted_items:
            ws.cell(row=row, column=1, value=value)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%")
            row += 1
        
        row += 1
        return row
    
    def _create_info_sheet(self, ws, result: FofaResult):
        """创建查询信息工作表"""
        row = 1
        
        # 标题
        ws.cell(row=row, column=1, value="查询详细信息")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        # 查询信息
        info_data = [
            ("查询模式", result.mode),
            ("查询页码", result.page),
            ("每页大小", result.size),
            ("总结果数", result.total),
            ("返回数量", len(result.results)),
            ("查询语句", result.query),
            ("导出时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        
        for label, value in info_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 80


# 测试代码
if __name__ == "__main__":
    from fofa_client import FofaResult
    
    # 创建测试数据
    test_result = FofaResult(
        mode='normal',
        page=1,
        size=5,
        total=100,
        results=[
            {
                'ip': '192.168.1.1',
                'port': '80',
                'protocol': 'http',
                'country': 'CN',
                'region': 'Beijing',
                'city': 'Beijing',
                'title': 'Test Server',
                'server': 'nginx',
            },
            {
                'ip': '192.168.1.2',
                'port': '443',
                'protocol': 'https',
                'country': 'CN',
                'region': 'Shanghai',
                'city': 'Shanghai',
                'title': 'Secure Server',
                'server': 'apache',
            },
        ],
        fields=['ip', 'port', 'protocol', 'country', 'region', 'city', 'title', 'server'],
        query='port="80" || port="443"'
    )
    
    # 测试导出
    manager = ExportManager(output_dir='./test_output')
    
    print("测试导出 Excel...")
    excel_path = manager.export_excel(test_result)
    print(f"Excel 文件已保存: {excel_path}")
    
    print("\n测试导出 CSV...")
    csv_path = manager.export_csv(test_result)
    print(f"CSV 文件已保存: {csv_path}")
    
    print("\n测试同时导出两种格式...")
    paths = manager.export_both(test_result)
    print(f"文件已保存: {paths}")
